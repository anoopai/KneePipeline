import os
import sys
import time
import shutil
import json
import random
import traceback
from weakref import ref
import numpy as np
import pandas as pd
from pathlib import Path
from statistics import mean
import torch
import os
from openpyxl import load_workbook
import pyvista as pv
from collections import defaultdict
import vtk
from vtk.util.numpy_support import numpy_to_vtk, vtk_to_numpy
from NSM.models import TriplanarDecoder
from NSM.mesh import create_mesh
from pymskt.mesh.meshTools import get_cartilage_properties_at_points
from pymskt.mesh import Mesh, BoneMesh, CartilageMesh
from NSM.mesh.interpolate import interpolate_points, interpolate_mesh, interpolate_common
from pymskt import mesh

def load_model(config, path_model_state, model_type='triplanar'):

    if model_type == 'triplanar':
        model_class = TriplanarDecoder
        params = {
            'latent_dim': config['latent_size'],
            'n_objects': config['objects_per_decoder'],
            'conv_hidden_dims': config['conv_hidden_dims'],
            'conv_deep_image_size': config['conv_deep_image_size'],
            'conv_norm': config['conv_norm'], 
            'conv_norm_type': config['conv_norm_type'],
            'conv_start_with_mlp': config['conv_start_with_mlp'],
            'sdf_latent_size': config['sdf_latent_size'],
            'sdf_hidden_dims': config['sdf_hidden_dims'],
            'sdf_weight_norm': config['weight_norm'],
            'sdf_final_activation': config['final_activation'],
            'sdf_activation': config['activation'],
            'sdf_dropout_prob': config['dropout_prob'],
            'sum_sdf_features': config['sum_conv_output_features'],
            'conv_pred_sdf': config['conv_pred_sdf'],
        }
    elif model_type == 'deepsdf':
        model_class = Decoder
        params = {
            'latent_size': config['latent_size'],
            'dims': config['layer_dimensions'],
            'dropout': config['layers_with_dropout'],
            'dropout_prob': config['dropout_prob'],
            'norm_layers': config['layers_with_norm'],
            'latent_in': config['layer_latent_in'],
            'weight_norm': config['weight_norm'],
            'xyz_in_all': config['xyz_in_all'],
            'latent_dropout': config['latent_dropout'],
            'activation': config['activation'],
            'final_activation': config['final_activation'],
            'concat_latent_input': config['concat_latent_input'],
            'n_objects': config['objects_per_decoder'],
            'progressive_add_depth': config['progressive_add_depth'],
            'layer_split': config['layer_split'],
        }
    else:
        raise ValueError(f'Unknown model type: {model_type}')


    model = model_class(**params)
    saved_model_state = torch.load(path_model_state)
    model.load_state_dict(saved_model_state["model"])
    model = model.cuda()
    model.eval()
    return model

def create_mean_mesh(latent_vector, model, model_type):

    latent_vector = torch.from_numpy(latent_vector).float().cuda()

    # CREATE MESHES & SAVE THEM TO DISK
    if model_type == 'bone_only':
        bone_mesh = create_mesh(decoder=model, latent_vector=latent_vector, objects=n_objects)
        bone_mesh.smooth
        n_sample = 10000
        bone_mesh.resample_surface(clusters=n_sample)
        return bone_mesh
        
    elif model_type == 'bone_cartilage':
        bone_mesh, cart_mesh = create_mesh(decoder=model, latent_vector=latent_vector, objects=n_objects)
        bone_mesh.smooth
        cart_mesh.smooth
        n_sample = 10000
        bone_mesh.resample_surface(clusters=n_sample)
        cart_mesh.resample_surface(clusters=n_sample)
        return bone_mesh, cart_mesh  
    else:
        raise ValueError(f'Input NSM model type')
    
def interpolate_matching_points(ref_latent, ref_mesh_path_femur, target_latent, model, model_type):
    if model_type == 'bone_only':
        ref_mesh = BoneMesh(ref_mesh_path_femur[0])
        ref_mesh_points= ref_mesh.points
        mesh_points_interp = interpolate_points(
            model= model,
            latent1= ref_latent,
            latent2= target_latent,
            n_steps= 100,
            points1= ref_mesh_points,
            surface_idx= 0,
            verbose= False,
            spherical= False
        )
        
        # interpolated_mesh= interpolate_mesh(
        #     model= model,
        #     latent1= ref_latent,
        #     latent2= target_latent,
        #     n_steps= 100,
        #     mesh= ref_mesh, 
        #     surface_idx= 0, 
        #     verbose= False, 
        #     spherical= True, 
        #     max_edge_len= 0.04, 
        #     adaptive= True, 
        #     smooth= True, 
        #     smooth_type= 'laplacian')
        
        # interpolated_mesh.resample_surface(clusters=10000)
        
        # create copies of the meshes, update the points, and save them to disk
        interpolated_mesh = ref_mesh.copy()
        interpolated_mesh.point_coords = mesh_points_interp
        return interpolated_mesh
        
    elif model_type == 'bone_cartilage':
        ref_mesh = BoneMesh(ref_mesh_path_femur[0])
        ref_mesh_points= ref_mesh.points
        mesh_points_interp_femur= interpolate_points(
            model=model,
            latent1= ref_latent,
            latent2= target_latent,
            n_steps= 100,
            points1= ref_mesh_points,
            surface_idx= 0,
            verbose= False,
            spherical= False
        )
        
        # interpolated_mesh_femur= interpolate_mesh(
        #     model= model,
        #     latent1= ref_latent,
        #     latent2= target_latent,
        #     n_steps= 100,
        #     mesh= ref_mesh, 
        #     surface_idx= 0, 
        #     verbose= False, 
        #     spherical= True, 
        #     max_edge_len= 0.04, 
        #     adaptive= True, 
        #     smooth= True, 
        #     smooth_type= 'laplacian')
        
        # interpolated_mesh_femur.resample_surface(clusters=10000)
        
        # interpolated_mesh_femur.copy_scalars_from_other_mesh_to_current(ref_mesh)
        
        # # create copies of the meshes, update the points, and save them to disk
        interpolated_mesh_femur= ref_mesh.copy()
        interpolated_mesh_femur.point_coords = mesh_points_interp_femur
        
        ref_mesh_cart = CartilageMesh(ref_mesh_path_femur[1])
        ref_mesh_points_cart= ref_mesh_cart.points
        mesh_points_interp_cart= interpolate_points(
            model=model,
            latent1= ref_latent,
            latent2= target_latent,
            n_steps= 100,
            points1= ref_mesh_points_cart,
            surface_idx= 1,
            verbose= False,
            spherical= False
        )
        
        # interpolated_mesh_cart= interpolate_mesh(
        #     model= model,
        #     latent1= ref_latent,
        #     latent2= target_latent,
        #     n_steps= 100,
        #     mesh= ref_mesh_cart, 
        #     surface_idx= 1, 
        #     verbose= False, 
        #     spherical= True, 
        #     max_edge_len= 0.04, 
        #     adaptive= True, 
        #     smooth= True, 
        #     smooth_type= 'laplacian')
        
        # interpolated_mesh_cart.resample_surface(clusters=10000)    
            
        # # create copies of the meshes, update the points, and save them to disk
        interpolated_mesh_cart= ref_mesh_cart.copy()
        interpolated_mesh_cart.point_coords = mesh_points_interp_cart
        
        return interpolated_mesh_femur, interpolated_mesh_cart
    else:
        raise ValueError(f'Unknown model type: {model_type}')

def compute_signed_distance_and_normal(mesh1_path, mesh2_path):
    """
    Compute signed distance from each point in mesh2 to the surface of mesh1.
    Signed distance is Positive outside, negative inside, zero on surface.
    Compute normal vector from mesh1 to mesh2.
    """
    mesh1 = pv.read(mesh1_path)
    mesh2 = pv.read(mesh2_path)
    
    assert mesh1.n_points == mesh2.n_points, "Resampling failed to produce matching number of points"

    # Build implicit distance function from surface of mesh1
    implicit_distance = vtk.vtkImplicitPolyDataDistance()
    surf = mesh1.extract_surface().triangulate().clean()
    implicit_distance.SetInput(surf)  # For older VTK versions

    # Evaluate signed distance at each point in mesh2
    signed_distances = np.array([
        implicit_distance.EvaluateFunction(point) for point in mesh2.points
    ])

    # Store result as point scalar
    mesh2.point_data["signed_distance"] = signed_distances
        
    # # Compute displacement vectors
    # disp_vectors = mesh2.points - mesh1.points  # shape (N, 3)
    # normx = disp_vectors[:, 0]
    # normy = disp_vectors[:, 1]
    # normz = disp_vectors[:, 2]
    # magnitude = np.linalg.norm(disp_vectors, axis=1)

    # # Attach as scalars to mesh2
    # mesh2.point_data['normx'] = normx
    # mesh2.point_data['normy'] = normy
    # mesh2.point_data['normz'] = normz
    
    # displacement_vectors = np.column_stack([
    # mesh2.point_data['normx'],
    # mesh2.point_data['normy'],
    # mesh2.point_data['normz']
    # ])
    # mesh2.point_data['displacement'] = displacement_vectors  # single vector field
    mesh = BoneMesh(mesh2)
    return mesh 

def get_icp_transform(model_path, bscore):
    
    '''
    This function generates mean mesh for a given Bscore along the Bscore vector from Mean healthy --> Mean OA
    '''
    
    import numpy as np
    import os
    import json

    path_json = os.path.join(model_path, 'model.json')

    with open(path_json, 'r') as f:
        model = json.load(f)
        
        bscore_vector = np.array(model['bscore_vector'])
        mean_healthy = np.array(model['mean_healthy'])
        std_healthy = np.array(model['std_healthy'])
        
        projection = (bscore * mean_healthy ) / std_healthy

        latent_vector = projection * bscore_vector
        
        return latent_vector

def interpolate_latent_for_bscore(model_path, bscore):
    """
    Interpolate latent vector for a given Bscore.
    """
    path_json = os.path.join(model_path, 'model.json')

    with open(path_json, 'r') as f:
        model = json.load(f)
        
        bscore_vector = np.array(model['bscore_vector'])
        mean_healthy = np.array(model['mean_healthy'])
        std_healthy = np.array(model['std_healthy'])
        
        # Compute norm squared of the bscore_vector
        bscore_vector_norm = np.linalg.norm(bscore_vector)

        # Your existing code
        projection = bscore * std_healthy + mean_healthy
        latent_vector = (projection / bscore_vector_norm) * bscore_vector
        
        # latent_vector = projection * bscore_vector
        
        return latent_vector

def project_fc_thickness_on_femur_mesh(bone_mesh_path, cart_mesh_path):

    bone_mesh= BoneMesh(bone_mesh_path)
    cart_mesh = CartilageMesh(cart_mesh_path)

    # IMPORTANT: The meshes should be "unnormailised" o patient size. Rough factor for scale-up is 45
    bone_mesh.point_coords = bone_mesh.point_coords * 45
    cart_mesh.point_coords = cart_mesh.point_coords * 45

    # setup the probe that we are using to get data from the T2 file 
    line_resolution = 100   # number of points along the line that the T2 data is sampled at
    no_thickness_filler = 0.0  # if no data is found, what value to fill the data with
    ray_cast_length= 10          # how far to extend the ray from the surface (using negative to go inwards/towards the other side)
    percent_ray_length_opposite_direction = 0.05  # extend the other way a % of the line to make sure get both edges. 1.0 = 100%|
    n_intersections = 2  # how many intersections to expect. If 2, then we expect to find two points on the ray that intersect the cartilage surface.

    thicknesses = np.zeros(bone_mesh.GetNumberOfPoints())

    node_data = get_cartilage_properties_at_points(
        bone_mesh,
        cart_mesh,
        t2_vtk_image=None,
        seg_vtk_image=None,
        ray_cast_length=ray_cast_length,
        percent_ray_length_opposite_direction=percent_ray_length_opposite_direction,
    )
    thicknesses += node_data

    # Assign the thickness scalars to the bone mesh surface.
    bone_mesh.point_data["thickness (mm)"] = thicknesses
    bone_mesh.point_coords = bone_mesh.point_coords / 45
    cart_mesh.point_coords = cart_mesh.point_coords / 45
    bone_mesh.save_mesh(bone_mesh_path)
    cart_mesh.save_mesh(cart_mesh_path)  
   
def calculate_fc_thickess_changes(bone_mesh_path_ref, bone_mesh_path_target):
    
    """
    Calculate the difference in cartilage thickness between two bone meshes.
    This function copies the cartilage thickness scalars from the reference mesh to the target mesh,
    calculates the difference in thickness, and saves the modified target mesh with the new scalars.
    Args:
        bone_mesh_path_ref (str): Path to the reference bone mesh file.
        bone_mesh_path_target (str): Path to the target bone mesh file.
    Returns:
        None: The function modifies the target mesh in place and saves it with the new scalars.
    
    """
    # Load the meshes
    bone_mesh1 = BoneMesh(bone_mesh_path_ref) 
    bone_mesh2 = BoneMesh(bone_mesh_path_target)
    
    bone_mesh2.copy_scalars_from_other_mesh_to_current(
        other_mesh=bone_mesh1,  # The mesh to copy scalars from
        orig_scalars_name='thickness (mm)',
        new_scalars_name= 'thickness (mm) Ref'
    )
    
    thickness_ref = vtk_to_numpy(bone_mesh2.GetPointData().GetArray('thickness (mm) Ref'))
    thickness = vtk_to_numpy(bone_mesh2.GetPointData().GetArray('thickness (mm)'))
    
    # Add edge detection - points with very low thickness are likely edge points
    edge_threshold = 0.25  # mm - adjust this value based on your data
    potential_edges = (thickness < edge_threshold) | (thickness_ref < edge_threshold)
    
    # Calculate the difference in thickness
    thickness_diff = thickness - thickness_ref
    
    # Set the difference to zero for edge points and where either thickness is zero
    thickness_diff[(thickness_ref == 0) | (thickness == 0) | potential_edges] = 0
    
    # Optional: Apply smoothing near edges to reduce artifacts
    # Find points near edges (within 2 points connectivity)
    near_edges = np.zeros_like(potential_edges)
    mesh_points = bone_mesh2.GetPoints()
    locator = vtk.vtkPointLocator()
    locator.SetDataSet(bone_mesh2)
    locator.BuildLocator()
    
    for i in range(len(potential_edges)):
        if potential_edges[i]:
            point = mesh_points.GetPoint(i)
            # Find points within a small radius
            id_list = vtk.vtkIdList()
            locator.FindPointsWithinRadius(0.5, point, id_list)  # 0.5mm radius
            for j in range(id_list.GetNumberOfIds()):
                near_edges[id_list.GetId(j)] = 1
    
    # Smooth the transition near edges
    for i in range(len(thickness_diff)):
        if near_edges[i] and not potential_edges[i]:
            # Find neighboring points
            id_list = vtk.vtkIdList()
            locator.FindClosestNPoints(5, mesh_points.GetPoint(i), id_list)
            # Calculate weighted average excluding edge points
            valid_diffs = []
            for j in range(id_list.GetNumberOfIds()):
                idx = id_list.GetId(j)
                if not potential_edges[idx]:
                    valid_diffs.append(thickness_diff[idx])
            if valid_diffs:
                thickness_diff[i] = np.mean(valid_diffs)
    
    # Add the difference scalar to the mesh
    bone_mesh2['thickness_diff (mm)'] = thickness_diff
    
    # Save the modified mesh
    bone_mesh2.save_mesh(bone_mesh_path_target)

def get_latent_for_subject_knee_visit(subject_id, knee_type, visit_num, results_all):
    """Retrieve the latent vector for a given subject, knee type, and visit number."""
    # Iterate through results to find the matching subject, knee, and visit
    for result in results_all:
        if (result['Subject'] == subject_id and 
            result['Knee'] == knee_type and 
            result['Visit'] == visit_num):
            return result['Latent']  # Return the latent vector if found
    return None  # Return None if no match is found

# Gather all the latent vectors
dir_path= '/dataNAS/people/anoopai/DESS_ACL_study'
data_path= os.path.join(dir_path, 'data')
log_path = '/dataNAS/people/anoopai/KneePipeline/logs'
output_file = os.path.join(log_path, f'shape_change_w.r.t_visit-1_subjectwise.txt')
path_config= '/dataNAS/people/anoopai/KneePipeline/config.json'

with open(path_config) as f:
    config = json.load(f)
    
model_types = ['bone_cartilage']
change_types = ['largest', 'smallest']

model_type_names= {
    'bone_only': 'BO',
    'bone_cartilage': 'BC'
}
knees_to_use = ['aclr', 'clat', 'ctrl']

try:
    for model_type in model_types:
         for change_type in change_types:
            print(f'Processing model type: {model_type}')       
            
            data_results_path= os.path.join(dir_path, 'results/BScore_and_FC_metrics')
            save_dir_path = os.path.join(data_results_path, f'mean_shape_recon')
            save_path = os.path.join(save_dir_path, f'{model_type_names[model_type]}_model/MinMax_shape_change_w.r.t_visit-1_subjectwise')
            change_save_path = os.path.join(save_path, f'{change_type}')
            
            if not os.path.exists(save_dir_path):
                os.makedirs(save_dir_path)
            if not os.path.exists(save_path):
                os.makedirs(save_path)
            if not os.path.exists(change_save_path):
                os.makedirs(change_save_path)
            
            bscore_data_path = os.path.join(data_results_path, 'data_BScore_Diff_graft.xlsx')
            data = pd.read_excel(bscore_data_path, sheet_name='VisitDiff')
            
            # Initialize an empty DataFrame for results
            results = pd.DataFrame()

            # Group by Visit
            for visit, group in data.groupby('Visit'):
                if change_type == 'largest':
                    # Find the top for BCScore Diff
                    aclr_best = group[group['Knee'] == 'aclr'].nlargest(1, f'{model_type_names[model_type]}Score Diff')
                    clat_best = group[group['Knee'] == 'clat'].nlargest(1, f'{model_type_names[model_type]}Score Diff')
                    ctrl_best = group[group['Knee'] == 'ctrl'].nlargest(1, f'{model_type_names[model_type]}Score Diff')
                else:
                    # Find the bottom for BCScore Diff
                    aclr_best = group[group['Knee'] == 'aclr'].nsmallest(1, f'{model_type_names[model_type]}Score Diff')
                    clat_best = group[group['Knee'] == 'clat'].nsmallest(1, f'{model_type_names[model_type]}Score Diff')
                    ctrl_best = group[group['Knee'] == 'ctrl'].nsmallest(1, f'{model_type_names[model_type]}Score Diff')

                # Combine results in one DataFrame
                combined = pd.concat([aclr_best, clat_best, ctrl_best],
                                    axis=0)

                # Add Visit to result DataFrame
                results_all = pd.concat([results, combined])
            
            # Exclude Visit-1, 2, 3, 4
            results_all = results_all[~results_all['Visit'].isin([1, 2, 3, 4])]
            
            results = results_all.copy()
            
            # Iterate through each subject-knee combination in results_all
            for index, row in results_all.iterrows():
                subject = row['Subject']
                knee = row['Knee']
                
                # Find the corresponding Visit 1 rows in data
                visit1_rows = data[(data['Subject'] == subject) & 
                                (data['Knee'] == knee) & 
                                (data['Visit'] == 1)]

                # Append those rows to the combined_results DataFrame
                results = pd.concat([results, visit1_rows], ignore_index=True)
            
            # Reset the index in the results
            results.reset_index(drop=True, inplace=True)
            
            # Assuming these variables are defined
            results_save_path = os.path.join(save_path, f'MinMax_change_in_BScore.xlsx')

            # Define the sheet name based on your model_type and change_type
            sheet_name = f'{model_type}_{change_type}'
            
            if not os.path.exists(results_save_path):
                # If the file does not exist, write the DataFrame to a new Excel file
                results.to_excel(results_save_path, sheet_name=sheet_name, index=False) 
                print(f"Created new file and wrote to sheet: {sheet_name}")
            else:
                # If the file exists, check if the sheet exists
                workbook = load_workbook(results_save_path)
                
                if sheet_name in workbook.sheetnames:
                    # If the sheet exists, do nothing
                    print(f"Sheet '{sheet_name}' already exists. No changes made.")
                else:
                    # If the sheet does not exist, append to the Excel file
                    with pd.ExcelWriter(results_save_path, engine='openpyxl', mode='a') as writer:
                        results.to_excel(writer, sheet_name=sheet_name, index=False) 
                        print(f"Added new sheet: {sheet_name} to existing file.")
                        
            # add a column for latent vector
            results['Latent'] = pd.NA  # Initializes with pandas NA
            
            # Iterate through the DataFrame and append latent values
            for index, row in results.iterrows():
                subject = row['Subject']
                knee = row['Knee']
                visit = row['Visit']

                knee_path = os.path.join(data_path, f'{subject}/VISIT-{visit}/{knee}')
                
                print(knee_path)
                
                if os.path.isdir(knee_path):
    
                    path_image =  os.path.join(knee_path,  'scans/qdess')
                    results_path =  os.path.join(knee_path, 'results_nsm')
                    qdess_results_file = os.path.join(results_path,  'qdess_results.json')
                    if model_type == 'bone_only':                                
                        nsm_recon_file = os.path.join(results_path, 'NSM_bone_only_recon_params.json')
                    elif model_type == 'bone_cartilage':
                        nsm_recon_file = os.path.join(results_path, 'NSM_recon_params.json')
                    else:
                        raise ValueError(f'Unknown model type: {model_type}')
                    
                    # Split the path into components
                    sub_component = Path(results_path).parts[6]  # '11-P'
                    visit_component = Path(results_path).parts[7]  # 'VISIT-1'
                    knee_component = Path(results_path).parts[8]  # 'clat'
                    
                    print('Subject:', sub_component, 'Visit:', visit_component, 'Knee:', knee_component)                                
                    if os.path.exists(nsm_recon_file):
                        
                        with open(nsm_recon_file) as f:
                            nsm_recon = json.load(f)
                        
                    
                    results.at[index, 'Latent'] = nsm_recon['latent']
        
            ########################## Compute meshes from NSM model and latents and register to mean mesh from Visit-1 ##########################
            
            print('Computing mean meshes from NSM model and latents')
            if model_type == 'bone_cartilage':

                # Setup file paths - relative to this script
                path_model_config = config['nsm']['path_model_config']
                path_model_state = config['nsm']['path_model_state']

                # Get BScore model path information
                path_model_folder = config['bscore']['path_model_folder'] 
                
                n_objects = 2
                
                # LOAD CONFIG
                with open(path_model_config, 'r') as f:
                    model_config = json.load(f)
            
                # Load Model
                model = load_model(model_config, path_model_state, model_type='triplanar')

            elif model_type == 'bone_only':
                # Setup file paths - relative to this script
                path_model_config = config['nsm_bone_only']['path_model_config']
                path_model_state = config['nsm_bone_only']['path_model_state']

                # Get BScore model path information
                path_model_folder = config['bscore_bone_only']['path_model_folder'] 
                
                n_objects = 1
            
                # LOAD CONFIG
                with open(path_model_config, 'r') as f:
                    model_config = json.load(f)
                
                # Load Model
                model = load_model(model_config, path_model_state, model_type='triplanar')

            else:
                raise ValueError(f'Unknown model type: {model_type}')
        

            # Create the mean mesh for Visit-1
            results_visit1= results[(results['Visit'] == 1)]
            for subject, knee, visit, latent in results[['Subject', 'Knee', 'Visit', 'Latent']].values:
                ref_latent_vector = np.array(latent).squeeze().astype(np.float64)

                print(f'Creating mesh for Visit-1 for {subject} {knee}')

                if model_type == 'bone_only':
                    ref_mesh_path_femur =os.path.join(change_save_path, f'Femur_NSMrecon_{subject}_{knee}_Visit-1.vtk')
                    if not os.path.exists(ref_mesh_path_femur):
                        bone_mesh= create_mean_mesh(ref_latent_vector, model, model_type)
                        bone_mesh.save_mesh(ref_mesh_path_femur)                
                elif model_type == 'bone_cartilage':                    
                    ref_mesh_path_femur =os.path.join(change_save_path, f'Femur_NSMrecon_{subject}_{knee}_Visit-1.vtk')
                    ref_mesh_path_cart =os.path.join(change_save_path, f'Cart_NSMrecon_{subject}_{knee}_Visit-1.vtk')
                    if not os.path.exists(ref_mesh_path_femur) or not os.path.exists(ref_mesh_path_cart):
                        bone_mesh, cart_mesh= create_mean_mesh(ref_latent_vector, model, model_type)
                        bone_mesh.save_mesh(ref_mesh_path_femur)
                        cart_mesh.save_mesh(ref_mesh_path_cart)
                        project_fc_thickness_on_femur_mesh(ref_mesh_path_femur, ref_mesh_path_cart)                                                
                else:
                    raise ValueError(f'Unknown model type: {model_type}')
                
            # Now we have the meshes for Visit-1, we can interpolate to Visit-5 
            results_visit5= results[(results['Visit'] == 5)]
            for subject, knee, visit, latent in results_visit5[['Subject', 'Knee', 'Visit', 'Latent']].values:

                target_latent = np.array(latent).squeeze().astype(np.float64)
                
                ref_latent = np.array(results_visit1[
                    (results_visit1['Subject'] == subject) & 
                    (results_visit1['Knee'] == knee) & 
                    (results_visit1['Visit'] == 1)  # explicitly looking for Visit 1
                ]['Latent'].values[0]).astype(np.float64)

                print(f'Processing {subject} {knee}')
                ref_mesh_path_femur = os.path.join(change_save_path, f'Femur_NSMrecon_{subject}_{knee}_Visit-1.vtk')
                ref_mesh_path_cart = os.path.join(change_save_path, f'Cart_NSMrecon_{subject}_{knee}_Visit-1.vtk')
                # target_mesh_path_femur = os.path.join(change_save_path, f'Femur_NSMrecon_{subject}_{knee}_Visit-5.vtk')
                # target_mesh_path_cart = os.path.join(change_save_path, f'Cart_NSMrecon_{subject}_{knee}_Visit-5.vtk')
            
                if model_type == 'bone_only':
                    interp_mesh= interpolate_matching_points(ref_latent, [ref_mesh_path_femur], target_latent, model, model_type)
                    interp_mesh_path =os.path.join(change_save_path, f'Femur_NSMrecon_{subject}_{knee}_Visit-1_to_Visit-5.vtk')
                    interp_mesh.save_mesh(interp_mesh_path)
                    deformed_mesh = compute_signed_distance_and_normal(ref_mesh_path_femur, interp_mesh_path)
                    deformed_mesh.save_mesh(interp_mesh_path)
            
                elif model_type == 'bone_cartilage':
                    interp_mesh_femur, interp_mesh_cart = interpolate_matching_points(ref_latent, [ref_mesh_path_femur, ref_mesh_path_cart], target_latent, model, model_type)
                    
                    interp_mesh_path_femur =os.path.join(change_save_path, f'Femur_NSMrecon_{subject}_{knee}_Visit-1_to_Visit-5.vtk')    
                    interp_mesh_path_cart =os.path.join(change_save_path, f'Cart_NSMrecon_{subject}_{knee}_Visit-1_to_Visit-5.vtk')
                    interp_mesh_femur.save_mesh(interp_mesh_path_femur)
                    interp_mesh_cart.save_mesh(interp_mesh_path_cart)                   
                    project_fc_thickness_on_femur_mesh(interp_mesh_path_femur, interp_mesh_path_cart)                                         
                    deformed_mesh_femur = compute_signed_distance_and_normal(ref_mesh_path_femur, interp_mesh_path_femur)
                    deformed_mesh_cart = compute_signed_distance_and_normal(ref_mesh_path_cart, interp_mesh_path_cart)
                    deformed_mesh_femur.save_mesh(interp_mesh_path_femur)
                    deformed_mesh_cart.save_mesh(interp_mesh_path_cart)
                    calculate_fc_thickess_changes(ref_mesh_path_femur, interp_mesh_path_femur)
                            
                else:
                    raise ValueError(f'Unknown model type: {model_type}')

except Exception as e:
    # Print the error message
    print("An error occurred:", e)
    # Print the traceback
    traceback.print_exc()
    # Write the error message to the output file
    with open(output_file, 'a') as f:
        f.write(f"An error occurred: {e}\n")